import requests
from bs4 import BeautifulSoup
import openpyxl

keyword = input("검색어를 입려해주세요 : ")

# load_workbooke() 기존의 엑셀파일 이후 데이터가 누적되어 저장
# 만약 파일이 없다면 오류가 발생
# wb = openpyxl.load_workbook("navernews.xlsx")
# wb = openpyxl.Workbook()

try :
    wb = openpyxl.load_workbook("navernews.xlsx")
    sheet = wb.active
except:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["기사제목", "언론사"])

# 컨테이너 : ul
# 기사제목 : a._sp_each_title
# 언론사 : span._sp_each_source

# 여러 페이지에서 수집되는 수집기
for n in range(1, 100, 10) :
    raw = requests.get(
        "https://search.naver.com/search.naver?where=news&sm=tab_jum&query="+keyword+"&start=" + str(n),
        headers={
            'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.103 Safari/537.36'})
    html = BeautifulSoup(raw.text, 'html.parser')
    articles = html.select("ul.type01 > li")

    for ar in articles:
        title = ar.select_one('a._sp_each_title').text
        source = ar.select_one('span._sp_each_source').text

        sheet.append([title, source])


wb.save("navernews.xlsx")